#!usr/bin/python3
import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

total_dict = dict()

row_id = 1
for row in ws:
        if row_id != 1:
                total = ws.cell(row = row_id, column = 3).value * 0.3
                total += ws.cell(row = row_id, column = 4).value * 0.35
                total += ws.cell(row = row_id, column = 5).value * 0.34
                total += ws.cell(row = row_id, column = 6).value
                ws.cell(row = row_id, column = 7).value = total
                #id = ws.cell(row = row_id, column = 1).value
                total_dict[row_id] = total
        row_id += 1


wb.save("student.xlsx")

len = len(total_dict)  #학생 수

#내림차순 정렬
sorted_list = sorted(total_dict.items(), key = lambda item: item[1], reverse=True)

#정렬된 순서대로 key list 저장
key = []
for i in range(len):
        key.append(sorted_list[i][0])
#print(key)

#등수 매기기
rank = dict()
for i in range(len):
        if total_dict[key[i]] == total_dict[key[i-1]]: #동점인 경우
                rank[key[i]] = rank[key[i-1]]
        else:
                rank[key[i]] = i + 1
#print(rank)

#rank 딕셔너리 value 리스트로 만들기
rank = list(rank.values())

#grade별 최대 등수
maxRank = []
maxRank.append(int(len * 0.3 * 0.5))  #A+
maxRank.append(int(len * 0.3))         #A0
maxRank.append(int(maxRank[1] + (len * 0.7 - maxRank[1]) * 0.5))  #B+
maxRank.append(int(len * 0.7))    #B0
maxRank.append(int(maxRank[3] + (len - maxRank[3]) * 0.5)) #C+
maxRank.append(int(len)) #C0

#학점 계산
grade_dic = dict()

for i in range(len):              
  if rank[i] <= maxRank[0]:  
    if rank.count(rank[i]) <= maxRank[0]:
      grade_dic[key[i]] = 'A+'               
    elif rank.count(rank[i]) <= maxRank[1]:
      grade_dic[key[i]] = 'A0'
    elif rank.count(rank[i]) <= maxRank[2]:
      grade_dic[key[i]] = 'B+'
    elif rank.count(rank[i]) <= maxRank[3]:
      grade_dic[key[i]] = 'B0'
    elif rank.count(rank[i]) <= maxRank[4]:
      grade_dic[key[i]] = 'C+'
    else:
      grade_dic[key[i]] = 'C0'

  elif rank[i] <= maxRank[1]:  
    if rank.count(rank[i]) <= maxRank[1]:
      grade_dic[key[i]] = 'A0'
    elif rank.count(rank[i]) <= maxRank[2]:
      grade_dic[key[i]] = 'B+'
    elif rank.count(rank[i]) <= maxRank[3]:
      grade_dic[key[i]] = 'B0'
    elif rank.count(rank[i]) <= maxRank[4]:
      grade_dic[key[i]] = 'C+'
    else:
      grade_dic[key[i]] = 'C0'

  elif rank[i] <= maxRank[2]:  
    if rank.count(rank[i]) <= maxRank[2]:
      grade_dic[key[i]] = 'B+'
    elif rank.count(rank[i]) <= maxRank[3]:
      grade_dic[key[i]] = 'B0'
    elif rank.count(rank[i]) <= maxRank[4]:
      grade_dic[key[i]] = 'C+'
    else:
      grade_dic[key[i]] = 'C0'

  elif rank[i] <= maxRank[3]:  
    if rank.count(rank[i]) <= maxRank[3]:
      grade_dic[key[i]] = 'B0'
    elif rank.count(rank[i]) <= maxRank[4]:
      grade_dic[key[i]] = 'C+'
    else:
      grade_dic[key[i]] = 'C0'

  elif rank[i] <= maxRank[4]:  
    if rank.count(rank[i]) <= maxRank[4]:
      grade_dic[key[i]] = 'C+'
    else:
      grade_dic[key[i]] = 'C0'
      
  else:
    grade_dic[key[i]] = 'C0'

print(grade_dic)

#엑셀 파일에 학점 반영
wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

row_id = 1
for row in ws:
        if row_id != 1:
                ws.cell(row = row_id, column = 8).value = grade_dic[row_id]
        row_id += 1
wb.save("student.xlsx")

